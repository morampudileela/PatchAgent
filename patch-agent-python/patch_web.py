#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Patch Agent  v2.3  -  Web UI
-----------------------------------------------------------------
Launch:
    python3 patch_web.py
Then open:  http://localhost:5000

Config:     config.yaml          (SSH credentials + LDAP settings)
Servers:    servers_template.xlsx  (cluster / server / service list)

v2.3 changes:
  - Login page: users authenticate with their own AD credentials
  - LDAP/AD group validation (two environments: nonprod / prod)
  - Server-side session store — credentials never sent to browser
  - SSH connections use the logged-in user's credentials
  - Sessions expire after 8 hours (matching password rotation policy)

v2.2 changes:
  - Group column: optional "Group" column scopes round-robin dependency
    checks to the same sub-tier

v2.1 changes:
  - Status column, dependency check, row selection fix, RR delay
"""

import functools
import json
import logging
import os
import queue
import threading
import time
import uuid
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import yaml
import openpyxl
from flask import (Flask, Response, g, jsonify, redirect,
                   render_template, request, url_for)

# -- optional paramiko ----------------------------------------------------
try:
    import paramiko
    _PARAMIKO = True
except ImportError:
    _PARAMIKO = False

# -- optional ldap3 -------------------------------------------------------
try:
    from ldap3 import Server as LdapServer, Connection, NTLM, SUBTREE, ALL
    _LDAP3 = True
except ImportError:
    _LDAP3 = False

# ========================================================================
#  App bootstrap
# ========================================================================

BASE_DIR    = Path(__file__).parent
SHARED_DIR  = BASE_DIR.parent          # SshTool/ — shared data files live here
CONFIG_PATH = BASE_DIR   / "config.yaml"
EXCEL_PATH  = SHARED_DIR / "servers_template.xlsx"
STATE_PATH  = SHARED_DIR / "patch_state.json"
LOG_PATH    = SHARED_DIR / "patch_agent.log"

app = Flask(__name__, template_folder=str(BASE_DIR / "templates"))
app.secret_key = os.environ.get("PATCH_AGENT_SECRET", "change-me-in-production-use-env-var")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("patch_web")

# -- in-memory job registry -----------------------------------------------
_jobs: dict[str, dict] = {}
_jobs_lock = threading.Lock()


# ========================================================================
#  Server-side session store
#  Credentials are stored in-memory on the server keyed by a random token.
#  Only the token is sent to the browser (HttpOnly cookie).
#  Sessions expire after SESSION_TTL seconds (= password rotation window).
# ========================================================================

SESSION_COOKIE = "pa_session"
SESSION_TTL    = 8 * 3600          # 8 hours

_sessions: dict[str, dict] = {}   # token -> {username, password, created}
_sessions_lock = threading.Lock()


def _create_session(username: str, password: str, environment: str = "nonprod") -> str:
    token = str(uuid.uuid4())
    now   = time.time()
    with _sessions_lock:
        # Prune expired sessions while we're holding the lock
        expired = [k for k, v in _sessions.items()
                   if now - v["created"] > SESSION_TTL]
        for k in expired:
            del _sessions[k]
        _sessions[token] = {"username":    username,
                             "password":    password,
                             "environment": environment,
                             "created":     now}
    return token


def _get_session(token: str | None) -> dict | None:
    if not token:
        return None
    with _sessions_lock:
        sess = _sessions.get(token)
        if not sess:
            return None
        if time.time() - sess["created"] > SESSION_TTL:
            del _sessions[token]
            return None
        return sess


def _destroy_session(token: str | None):
    if token:
        with _sessions_lock:
            _sessions.pop(token, None)


def login_required(f):
    """Decorator that enforces authentication for routes."""
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        token = request.cookies.get(SESSION_COOKIE)
        sess  = _get_session(token)
        if not sess:
            if request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized",
                                "login_required": True}), 401
            return redirect(url_for("login"))
        g.current_user = sess   # {username, password, created}
        return f(*args, **kwargs)
    return decorated


# ========================================================================
#  Config
# ========================================================================

def load_config() -> dict:
    if not CONFIG_PATH.exists():
        return {"ssh": {}, "patching": {}, "ldap": {}}
    with open(CONFIG_PATH) as f:
        raw = yaml.safe_load(f)
    if not isinstance(raw, dict):
        return {"ssh": {}, "patching": {}, "ldap": {}}
    for key in ("ssh", "patching", "ldap"):
        if not isinstance(raw.get(key), dict):
            raw[key] = {}
    return raw


def _load_ldap_cfg(environment: str | None = None) -> dict:
    """Return the LDAP config block for the given environment.
    Falls back to ldap.environment in config.yaml when not supplied."""
    ldap = load_config().get("ldap", {})
    env  = environment if environment in ("nonprod", "prod") \
           else ldap.get("environment", "nonprod")
    return ldap.get(env, {})


# ========================================================================
#  LDAP / Active Directory authentication
# ========================================================================

def _ldap_authenticate(username: str, password: str,
                       environment: str = "nonprod") -> tuple[bool, str]:
    """
    Validate username + password against Active Directory for the given environment.
    Checks group membership if required_group is configured.
    Returns (success, error_message).
    """
    if not _LDAP3:
        logger.error("ldap3 is not installed — run: pip install ldap3")
        return False, "LDAP library not installed on server"

    cfg            = _load_ldap_cfg(environment)
    server_url     = cfg.get("server",         "ldap://dc01.company.com:389")
    domain         = cfg.get("domain",         "COMPANY")
    base_dn        = cfg.get("base_dn",        "DC=company,DC=com")
    required_group = cfg.get("required_group", "")

    if not username or not password:
        return False, "Username and password are required"

    try:
        server  = LdapServer(server_url, get_info=ALL, connect_timeout=5)
        user_dn = f"{domain}\\{username}"
        conn    = Connection(server, user=user_dn, password=password,
                             authentication=NTLM, auto_bind=False)

        if not conn.bind():
            logger.warning("LDAP bind failed for user: %s", username)
            return False, "Invalid username or password"

        # Group membership check (uses LDAP_MATCHING_RULE_IN_CHAIN for
        # recursive/nested group resolution)
        if required_group:
            conn.search(
                search_base   = base_dn,
                search_filter = (
                    f"(&(sAMAccountName={username})"
                    f"(memberOf:1.2.840.113556.1.4.1941:={required_group}))"
                ),
                search_scope  = SUBTREE,
                attributes    = ["sAMAccountName"],
            )
            if not conn.entries:
                conn.unbind()
                logger.warning("User %s is not in required group: %s",
                               username, required_group)
                return False, "Access denied: not a member of the required group"

        conn.unbind()
        logger.info("LDAP authentication successful for user: %s", username)
        return True, ""

    except Exception as exc:
        logger.error("LDAP error for user %s: %s", username, exc)
        return False, f"LDAP connection error — please try again"


# ========================================================================
#  Excel loading
# ========================================================================

COL_MAP = {
    "id":                             "id",
    "cluster":                        "cluster",
    "server name":                    "server_name",
    "server":                         "server_name",
    "hostname / ip":                  "host",
    "hostname/ip":                    "host",
    "hostname":                       "host",
    "ip":                             "host",
    "host":                           "host",
    "service name":                   "service",
    "service":                        "service",
    "stop command":                   "stop_cmd",
    "stop":                           "stop_cmd",
    "start command":                  "start_cmd",
    "start":                          "start_cmd",
    "mode":                           "mode",
    "mode\n(round_robin / batch)":    "mode",
    "delay":                          "rr_delay",
    "delay (s)":                      "rr_delay",
    "delay(s)":                       "rr_delay",
    "rr delay":                       "rr_delay",
    "round robin delay":              "rr_delay",
    "status check command":           "status_cmd",
    "status check":                   "status_cmd",
    "status cmd":                     "status_cmd",
    "check command":                  "status_cmd",
    "group":                          "group",
    "server group":                   "group",
    "tier":                           "group",
    "service group":                  "group",
    "notes":                          "notes",
    "environment":                    "environment",
    "env":                            "environment",
}

REQUIRED = {"id", "cluster", "host", "service", "stop_cmd", "start_cmd", "mode"}


def load_servers(excel_path: Path = EXCEL_PATH) -> list[dict]:
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws_name = next((s for s in wb.sheetnames if s.lower() != "legend"), wb.sheetnames[0])
    ws = wb[ws_name]

    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        raise ValueError("Excel sheet is empty")

    raw_hdrs = [str(h).strip() if h else "" for h in raw_rows[0]]
    headers  = [COL_MAP.get(h.lower(), h.lower()) for h in raw_hdrs]

    missing = REQUIRED - set(headers)
    if missing:
        raise ValueError(f"Excel missing columns: {missing}")

    servers = []
    for row in raw_rows[1:]:
        rec = dict(zip(headers, row))
        if not rec.get("host"):
            continue
        try:
            rec["id"] = int(rec.get("id") or 0)
        except (TypeError, ValueError):
            rec["id"] = 0
        mode = str(rec.get("mode") or "batch").strip().lower()
        if mode not in ("round_robin", "batch"):
            mode = "batch"
        rec["mode"]        = mode
        rec["cluster"]     = str(rec.get("cluster") or "Default").strip()
        rec["server_name"] = str(rec.get("server_name") or rec["host"]).strip()
        rec["service"]     = str(rec.get("service") or "service").strip()
        rec["stop_cmd"]    = str(rec.get("stop_cmd") or "").strip()
        rec["start_cmd"]   = str(rec.get("start_cmd") or "").strip()
        rec["status_cmd"]  = str(rec.get("status_cmd") or "").strip()
        rec["group"]       = str(rec.get("group") or "").strip()
        rec["notes"]       = str(rec.get("notes") or "").strip()
        raw_delay = rec.get("rr_delay")
        try:
            rec["rr_delay"] = float(raw_delay) if raw_delay not in (None, "", "None") else None
        except (TypeError, ValueError):
            rec["rr_delay"] = None
        servers.append(rec)

    return servers


def parse_row_selection(selection: str, all_ids: list[int]) -> list[int]:
    selection = selection.strip()
    if not selection or selection == "*":
        return sorted(all_ids)

    id_set = set()
    for part in selection.split(","):
        part = part.strip()
        if "-" in part:
            lo, _, hi = part.partition("-")
            try:
                id_set.update(range(int(lo.strip()), int(hi.strip()) + 1))
            except ValueError:
                pass
        else:
            try:
                id_set.add(int(part))
            except ValueError:
                pass

    valid = set(all_ids)
    return sorted(id_set & valid)


# ========================================================================
#  SSH helpers
# ========================================================================

def _ssh_connect(host: str, cfg: dict,
                 username: str | None = None,
                 password: str | None = None):
    """
    Open an SSH connection to host.
    username / password come from the logged-in user's session.
    Falls back to config.yaml credentials if not provided (e.g. key-based auth).
    """
    ssh_cfg = cfg.get("ssh") or {}
    c = paramiko.SSHClient()
    c.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    kw = dict(
        hostname       = host,
        port           = int(ssh_cfg.get("port", 22)),
        username       = username or ssh_cfg.get("username", "root"),
        timeout        = float(ssh_cfg.get("connect_timeout", 15)),
        banner_timeout = 30,
        auth_timeout   = 30,
    )
    if password:
        kw["password"] = password
    else:
        key_path = (ssh_cfg.get("private_key_path") or "").strip()
        if key_path:
            kw["key_filename"] = os.path.expanduser(key_path)
        else:
            fallback_pw = (ssh_cfg.get("password") or "").strip()
            if fallback_pw:
                kw["password"] = fallback_pw
    c.connect(**kw)
    return c


def _run(client, cmd: str, timeout: float) -> tuple[int, str, str]:
    _, out, err = client.exec_command(cmd, timeout=timeout)
    code = out.channel.recv_exit_status()
    return code, out.read().decode().strip(), err.read().decode().strip()


def _check_status(row: dict, cfg: dict,
                  username: str | None = None,
                  password: str | None = None) -> str:
    if not row.get("status_cmd"):
        return "unknown"
    if not _PARAMIKO:
        return "unknown"
    try:
        client = _ssh_connect(row["host"], cfg,
                               username=username, password=password)
        code, _, _ = _run(client, row["status_cmd"], 10.0)
        client.close()
        return "running" if code == 0 else "stopped"
    except Exception:
        return "error"


# ========================================================================
#  Job execution
# ========================================================================

def _emit(q: queue.Queue, level: str, message: str, extra: dict | None = None):
    ev = {
        "ts":      datetime.now(timezone.utc).strftime("%H:%M:%S"),
        "level":   level,
        "message": message,
    }
    if extra:
        ev.update(extra)
    q.put(ev)
    getattr(logger, "info" if level in ("ok", "info") else level, logger.info)(message)


def execute_service(row: dict, action: str, cfg: dict,
                    dry_run: bool, q: queue.Queue,
                    username: str | None = None,
                    password: str | None = None) -> dict:
    host    = row["host"]
    service = row["service"]
    row_id  = row["id"]
    cmd     = row["stop_cmd"] if action == "stop" else row["start_cmd"]
    timeout = float((cfg.get("ssh") or {}).get("command_timeout", 30))
    label   = f"[{row_id}] {host} / {service}"

    result = {"id": row_id, "host": host, "service": service,
              "action": action, "status": "ok", "message": ""}

    if dry_run:
        _emit(q, "info", f"{label}  [DRY-RUN] Would {action}: {cmd}",
              {"row_id": row_id, "host": host, "service": service})
        result["message"] = f"Dry-run - would run: {cmd}"
        return result

    if not _PARAMIKO:
        msg = "paramiko not installed"
        _emit(q, "error", f"{label}  X  {msg}", {"row_id": row_id})
        result.update(status="error", message=msg)
        return result

    # Skip STOP if service already stopped
    if action == "stop" and row.get("status_cmd"):
        try:
            chk_client = _ssh_connect(host, cfg,
                                       username=username, password=password)
            chk_code, _, _ = _run(chk_client, row["status_cmd"], 10.0)
            chk_client.close()
            if chk_code != 0:
                _emit(q, "warn",
                      f"{label}  -- Already stopped, skipping",
                      {"row_id": row_id, "host": host, "service": service})
                result["message"] = "Already stopped - skipped"
                return result
        except Exception:
            pass

    try:
        client = _ssh_connect(host, cfg, username=username, password=password)
        code, out, err = _run(client, cmd, timeout)
        client.close()

        if code == 0:
            _emit(q, "ok",
                  f"{label}  OK  {'STOPPED' if action == 'stop' else 'STARTED'}",
                  {"row_id": row_id, "host": host, "service": service})
            result["message"] = f"Service {action}ped successfully"
        else:
            err_msg = err or out or f"exit {code}"
            _emit(q, "error", f"{label}  X  Failed to {action}: {err_msg}",
                  {"row_id": row_id, "host": host, "service": service})
            result.update(status="error", message=f"Command failed: {err_msg}")

    except Exception as exc:
        _emit(q, "error", f"{label}  X  {exc}",
              {"row_id": row_id, "host": host, "service": service})
        result.update(status="error", message=str(exc))

    return result


def _process_server_group(group_rows: list[dict], action: str, cfg: dict,
                           dry_run: bool, q: queue.Queue,
                           username: str | None = None,
                           password: str | None = None) -> list[dict]:
    if action == "stop":
        ordered = sorted(group_rows, key=lambda r: r["id"])
    else:
        ordered = sorted(group_rows, key=lambda r: r["id"], reverse=True)

    results = []
    for row in ordered:
        results.append(execute_service(row, action, cfg, dry_run, q,
                                       username=username, password=password))
    return results


def run_job(job_id: str, rows: list[dict], action: str,
            cfg: dict, dry_run: bool,
            username: str | None = None,
            password: str | None = None):
    """
    Background thread: groups rows by server, runs round-robin then batch.
    Credentials (username/password) come from the user's session and are used
    for every SSH connection in this job.
    """
    with _jobs_lock:
        job = _jobs[job_id]
    q       = job["events_q"]
    results = []

    total_servers = 0

    try:
        server_groups: dict[tuple, list] = defaultdict(list)
        for row in rows:
            key = (row["cluster"], row["host"])
            server_groups[key].append(row)

        rr_groups    = [(k, v) for k, v in server_groups.items()
                        if v[0]["mode"] == "round_robin"]
        batch_groups = [(k, v) for k, v in server_groups.items()
                        if v[0]["mode"] == "batch"]

        # Build group lookup for dependency scoping
        server_group_map: dict[tuple, str] = {
            k: v[0].get("group", "") for k, v in server_groups.items()
        }

        total_servers = len(server_groups)
        _emit(q, "info",
              f"Job {job_id[:8]}  action={action.upper()}  "
              f"servers={total_servers} ({len(rr_groups)} round-robin, "
              f"{len(batch_groups)} batch)  dry_run={dry_run}  "
              f"user={username or 'config'}",
              {"progress": 0, "total": total_servers})

        done  = 0
        delay = float((cfg.get("patching") or {}).get("round_robin_delay", 5))

        # -- Round-robin phase ----------------------------------------
        if rr_groups:
            _emit(q, "info",
                  f"-- PHASE 1: Round-Robin  ({len(rr_groups)} servers) --")

            already_stopped: set[tuple] = set()
            if action == "stop" and not dry_run:
                _emit(q, "info", "  Pre-flight: checking round-robin server statuses ...")
                pf_lock = threading.Lock()

                def _pf_check(key, group_rows):
                    chk = next((r for r in group_rows if r.get("status_cmd")), None)
                    if chk:
                        st = _check_status(chk, cfg,
                                           username=username, password=password)
                        if st == "stopped":
                            with pf_lock:
                                already_stopped.add(key)
                            _emit(q, "warn",
                                  f"  [pre-flight] {key[0]}/{key[1]} is already STOPPED")

                pf_threads = [
                    threading.Thread(target=_pf_check, args=(k, v), daemon=True)
                    for k, v in rr_groups
                ]
                for t in pf_threads: t.start()
                for t in pf_threads: t.join(timeout=20)

                if already_stopped:
                    _emit(q, "warn",
                          f"  {len(already_stopped)} server(s) already stopped — "
                          f"dependency check active for remaining servers")

            we_stopped: set[tuple] = set()

            for idx, ((cluster, host), group_rows) in enumerate(rr_groups):
                key   = (cluster, host)
                sname = group_rows[0]["server_name"]

                if action == "stop" and not dry_run:
                    my_group = server_group_map.get(key, "")
                    blocking = [
                        f"{k[1]}" for k in already_stopped
                        if k != key
                        and k not in we_stopped
                        and my_group != ""
                        and server_group_map.get(k, "") == my_group
                    ]
                    if blocking:
                        _emit(q, "warn",
                              f"  SKIP [{idx+1}/{len(rr_groups)}] {cluster} / {sname}"
                              f" — other RR server(s) already stopped: {', '.join(blocking)}")
                        done += 1
                        pct = int(done / total_servers * 100)
                        _emit(q, "warn", f"  Progress: {done}/{total_servers}",
                              {"progress": pct, "total": total_servers})
                        continue

                _emit(q, "info",
                      f"  [{idx+1}/{len(rr_groups)}] {cluster} / {sname} ({host})")
                res = _process_server_group(group_rows, action, cfg, dry_run, q,
                                            username=username, password=password)
                results.extend(res)

                if action == "stop" and all(r["status"] == "ok" for r in res):
                    we_stopped.add(key)

                done += 1
                pct = int(done / total_servers * 100)
                _emit(q, "info", f"  Progress: {done}/{total_servers}",
                      {"progress": pct, "total": total_servers})

                if idx < len(rr_groups) - 1 and not dry_run:
                    server_delay = group_rows[0].get("rr_delay")
                    effective_delay = server_delay if server_delay is not None else delay
                    if effective_delay > 0:
                        _emit(q, "warn",
                              f"  ~~~ Waiting {int(effective_delay)}s before next server ~~~")
                        time.sleep(effective_delay)

        # -- Batch phase ----------------------------------------------
        if batch_groups:
            _emit(q, "info",
                  f"-- PHASE 2: Batch  ({len(batch_groups)} servers, parallel) --")
            batch_results: list = [None] * len(batch_groups)
            threads = []
            lock    = threading.Lock()

            def _run_group(idx, cluster, host, group_rows):
                sname = group_rows[0]["server_name"]
                _emit(q, "info", f"  Starting {cluster} / {sname} ({host}) ...")
                res = _process_server_group(group_rows, action, cfg, dry_run, q,
                                            username=username, password=password)
                with lock:
                    batch_results[idx] = res
                    nonlocal done
                    done += 1
                    pct = int(done / total_servers * 100)
                    _emit(q, "info", f"  Progress: {done}/{total_servers}",
                          {"progress": pct, "total": total_servers})

            max_w = int((cfg.get("patching") or {}).get("batch_max_workers", 10))
            sem   = threading.Semaphore(max_w)

            def _worker(idx, cluster, host, group_rows):
                with sem:
                    _run_group(idx, cluster, host, group_rows)

            for i, ((cluster, host), group_rows) in enumerate(batch_groups):
                t = threading.Thread(target=_worker,
                                     args=(i, cluster, host, group_rows),
                                     daemon=True)
                threads.append(t)
                t.start()

            for t in threads:
                t.join()

            for res in batch_results:
                if res:
                    results.extend(res)

        # -- Done -----------------------------------------------------
        errors = sum(1 for r in results if r["status"] != "ok")
        _emit(q, "ok" if errors == 0 else "warn",
              f"-- COMPLETE  "
              f"processed={len(results)}  ok={len(results)-errors}  errors={errors} --",
              {"progress": 100, "total": total_servers, "done": True})

        _save_state(results, action)

        with _jobs_lock:
            _jobs[job_id]["status"]  = "done"
            _jobs[job_id]["results"] = results

    except Exception as exc:
        _emit(q, "error", f"Job crashed: {exc}", {"done": True})
        with _jobs_lock:
            _jobs[job_id]["status"] = "error"

    finally:
        q.put(None)


# ========================================================================
#  State persistence
# ========================================================================

def _save_state(results: list[dict], action: str):
    state: dict = {}
    if STATE_PATH.exists():
        try:
            state = json.loads(STATE_PATH.read_text())
        except Exception:
            state = {}

    ts = datetime.now(timezone.utc).isoformat()
    session = {
        "action":     action,
        "started_at": ts,
        "results":    results,
        "summary": {
            "total":  len(results),
            "ok":     sum(1 for r in results if r["status"] == "ok"),
            "errors": sum(1 for r in results if r["status"] != "ok"),
        },
    }
    state.setdefault("sessions", []).append(session)
    STATE_PATH.write_text(json.dumps(state, indent=2))


def _load_state() -> dict:
    if not STATE_PATH.exists():
        return {"sessions": []}
    try:
        return json.loads(STATE_PATH.read_text())
    except Exception:
        return {"sessions": []}


# ========================================================================
#  Flask routes
# ========================================================================

# -- Auth -----------------------------------------------------------------

@app.route("/login", methods=["GET", "POST"])
def login():
    # Already logged in — go to app
    token = request.cookies.get(SESSION_COOKIE)
    if _get_session(token):
        return redirect(url_for("index"))

    error    = None
    username = ""
    environment = "nonprod"
    if request.method == "POST":
        username    = (request.form.get("username") or "").strip()
        password    = request.form.get("password") or ""
        environment = request.form.get("environment", "nonprod")
        if environment not in ("nonprod", "prod"):
            environment = "nonprod"

        ok, msg = _ldap_authenticate(username, password, environment)
        if ok:
            token = _create_session(username, password, environment)
            resp  = redirect(url_for("index"))
            resp.set_cookie(SESSION_COOKIE, token,
                            httponly=True, samesite="Lax",
                            max_age=SESSION_TTL)
            return resp
        else:
            error = msg or "Authentication failed"

    return render_template("login.html",
                           error=error,
                           username=username,
                           environment=environment)


@app.route("/logout")
def logout():
    token = request.cookies.get(SESSION_COOKIE)
    _destroy_session(token)
    resp = redirect(url_for("login"))
    resp.delete_cookie(SESSION_COOKIE)
    return resp


# -- App ------------------------------------------------------------------

@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/api/me")
@login_required
def api_me():
    """Returns the currently logged-in username and environment."""
    return jsonify({
        "username":    g.current_user["username"],
        "environment": g.current_user.get("environment", "nonprod"),
    })


@app.route("/api/servers")
@login_required
def api_servers():
    try:
        env      = g.current_user.get("environment", "nonprod")
        all_rows = load_servers()
        # Filter to rows that match the session environment.
        # Rows without an environment column are included in all environments
        # (backwards compatibility with Excel files that predate this column).
        rows     = [r for r in all_rows
                    if not r.get("environment") or r["environment"] == env]
        clusters = sorted({r["cluster"] for r in rows})
        return jsonify({"rows": rows, "clusters": clusters})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/status", methods=["POST"])
@login_required
def api_status():
    body    = request.get_json(force=True) or {}
    row_ids = set(int(x) for x in body.get("row_ids", []))
    username = g.current_user["username"]
    password = g.current_user["password"]

    try:
        all_rows = load_servers()
        rows     = [r for r in all_rows if not row_ids or r["id"] in row_ids]
        cfg      = load_config()

        statuses    = {}
        status_lock = threading.Lock()

        def check_row(row):
            rid    = row["id"]
            status = _check_status(row, cfg,
                                   username=username, password=password)
            with status_lock:
                statuses[str(rid)] = status

        max_w = int((cfg.get("patching") or {}).get("batch_max_workers", 10))
        sem   = threading.Semaphore(max_w)
        threads = []

        def worker(row):
            with sem:
                check_row(row)

        for r in rows:
            t = threading.Thread(target=worker, args=(r,), daemon=True)
            threads.append(t)
            t.start()

        for t in threads:
            t.join(timeout=25)

        return jsonify({"statuses": statuses})

    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/job/start", methods=["POST"])
@login_required
def api_job_start():
    body    = request.get_json(force=True) or {}
    action  = body.get("action", "stop").lower()
    dry_run = bool(body.get("dry_run", False))
    sel_str = str(body.get("selection", "*"))

    # Capture credentials at request time (not in background thread)
    username = g.current_user["username"]
    password = g.current_user["password"]

    if action not in ("stop", "start"):
        return jsonify({"error": "action must be 'stop' or 'start'"}), 400

    try:
        all_rows  = load_servers()
        all_ids   = [r["id"] for r in all_rows]
        sel_ids   = parse_row_selection(sel_str, all_ids)
        sel_rows  = [r for r in all_rows if r["id"] in set(sel_ids)]
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    if not sel_rows:
        return jsonify({"error": "No rows matched the selection"}), 400

    cfg    = load_config()
    job_id = str(uuid.uuid4())

    with _jobs_lock:
        _jobs[job_id] = {
            "status":    "running",
            "action":    action,
            "dry_run":   dry_run,
            "selection": sel_str,
            "user":      username,
            "events_q":  queue.Queue(),
            "results":   [],
        }

    t = threading.Thread(
        target=run_job,
        args=(job_id, sel_rows, action, cfg, dry_run),
        kwargs={"username": username, "password": password},
        daemon=True,
    )
    t.start()

    return jsonify({"job_id": job_id, "server_count": len(sel_rows)})


@app.route("/api/job/stream/<job_id>")
@login_required
def api_job_stream(job_id: str):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        return jsonify({"error": "job not found"}), 404

    def _generate():
        q = job["events_q"]
        while True:
            try:
                ev = q.get(timeout=30)
            except queue.Empty:
                yield "event: heartbeat\ndata: {}\n\n"
                continue
            if ev is None:
                yield "event: done\ndata: {}\n\n"
                break
            yield f"data: {json.dumps(ev)}\n\n"

    return Response(_generate(),
                    mimetype="text/event-stream",
                    headers={"Cache-Control":    "no-cache",
                             "X-Accel-Buffering": "no"})


@app.route("/api/job/<job_id>")
@login_required
def api_job_status(job_id: str):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    return jsonify({
        "status":  job["status"],
        "action":  job["action"],
        "results": job["results"],
    })


@app.route("/api/history")
@login_required
def api_history():
    state    = _load_state()
    sessions = state.get("sessions", [])[-20:]
    sessions.reverse()
    return jsonify({"sessions": sessions})


@app.route("/api/resolve", methods=["POST"])
@login_required
def api_resolve():
    body = request.get_json(force=True) or {}
    sel  = str(body.get("selection", ""))
    try:
        rows    = load_servers()
        all_ids = [r["id"] for r in rows]
        sel_ids = parse_row_selection(sel, all_ids)
        return jsonify({"ids": sel_ids, "count": len(sel_ids)})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ========================================================================
#  Entry point
# ========================================================================

if __name__ == "__main__":
    print("\n  Patch Agent  v2.3")
    print(f"  Config:  {CONFIG_PATH}")
    print(f"  Excel:   {EXCEL_PATH}")
    print(f"  LDAP:    {_load_ldap_cfg().get('server', '(not configured)')}")
    print(f"  Open:    http://localhost:5000\n")
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
