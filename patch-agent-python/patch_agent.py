#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Linux Server Patching Agent  v1.1
------------------------------------------------------------
Reads an Excel file with server details, SSHs into each one,
checks the service status, then toggles it:
  • Running  → stop the service
  • Stopped  → start the service

Two execution modes (set per-row in the Excel):
  • round_robin – servers processed one at a time, sequentially
  • batch       – servers processed in parallel (threadpool)

Progress is saved to patch_state.json after every server so you
can resume a run that was interrupted or handed off mid-way.

Usage examples
--------------
  # Patch ALL servers
  python3 patch_agent.py

  # Patch a specific row range (e.g. hand off rows 1–8 to a colleague)
  python3 patch_agent.py --start-row 1 --stop-row 8
  python3 patch_agent.py --start-row 9 --stop-row 15

  # Resume — skip rows already marked completed in patch_state.json
  python3 patch_agent.py --resume

  # Resume within a range
  python3 patch_agent.py --start-row 5 --stop-row 10 --resume

  # Show current progress without running anything
  python3 patch_agent.py --status

  # Clear saved progress and start fresh
  python3 patch_agent.py --reset-state

  # Patch specific individual row IDs (comma-separated)
  python3 patch_agent.py --rows 1,3,5,7

  # Dry-run: show what would happen without touching servers
  python3 patch_agent.py --dry-run

  # Verbose SSH output
  python3 patch_agent.py --verbose
"""

import argparse
import json
import logging
import os
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

import yaml
import openpyxl

# -- Optional: colorama for coloured terminal output -----------------------
try:
    from colorama import init as colorama_init, Fore, Style
    colorama_init(autoreset=True)
    _COLOR = True
except ImportError:
    _COLOR = False

    class _Stub:
        def __getattr__(self, _):
            return ""

    Fore = Style = _Stub()

# -- Optional: paramiko for SSH --------------------------------------------
try:
    import paramiko
    _PARAMIKO = True
except ImportError:
    _PARAMIKO = False

# =========================================================================
#  Terminal helpers
# =========================================================================

BANNER = r"""
  ____       _       _        _                    _
 |  _ \ __ _| |_ ___| |__    / \   __ _  ___ _ __ | |_
 | |_) / _` | __/ __| '_ \  / _ \ / _` |/ _ \ '_ \| __|
 |  __/ (_| | || (__| | | |/ ___ \ (_| |  __/ | | | |_
 |_|   \__,_|\__\___|_| |_/_/   \_\__, |\___|_| |_|\__|
                                   |___/
  Linux Server Patching Agent  •  v1.1
"""

STATE_FILE = str(Path(__file__).parent.parent / "patch_state.json")


def _c(text: str, colour: str) -> str:
    return f"{colour}{text}{Style.RESET_ALL}" if _COLOR else text


def ok(msg):   print(_c(f"  ✔  {msg}", Fore.GREEN))
def warn(msg): print(_c(f"  ⚠  {msg}", Fore.YELLOW))
def err(msg):  print(_c(f"  ✖  {msg}", Fore.RED))
def info(msg): print(_c(f"  ·  {msg}", Fore.CYAN))
def hdr(msg):  print(_c(f"\n{'-'*60}\n  {msg}\n{'-'*60}", Fore.BLUE))


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def setup_logging(log_file: str) -> logging.Logger:
    logger = logging.getLogger("patch_agent")
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s",
                            datefmt="%Y-%m-%d %H:%M:%S")
    logger.addHandler(logging.StreamHandler(stream=open(os.devnull, "w")))
    if log_file:
        fh = logging.FileHandler(log_file)
        fh.setFormatter(fmt)
        logger.addHandler(fh)
    return logger


# =========================================================================
#  State file  (patch_state.json)
# =========================================================================

class PatchState:
    """
    Thread-safe progress tracker.  Persists after every server so that a
    run that is interrupted (Ctrl-C, power loss, handoff) can be resumed
    with --resume.

    Schema
    ------
    {
      "version":           "1.1",
      "excel":             "servers_template.xlsx",
      "started_at":        "2026-05-04T10:00:00Z",
      "last_updated":      "2026-05-04T10:07:43Z",
      "completed_row_ids": [1, 2, 3, 5],
      "failed_row_ids":    [4],
      "results": [ { id, host, service, action, status, message, completed_at }, … ]
    }
    """

    def __init__(self, excel: str, state_file: str = STATE_FILE):
        self._path  = Path(state_file)
        self._lock  = threading.Lock()
        self._excel = excel
        self._data: dict = {}
        self._load_or_init()

    # -- internal ----------------------------------------------------------

    def _load_or_init(self):
        if self._path.exists():
            try:
                with open(self._path) as f:
                    self._data = json.load(f)
                info(f"Loaded existing state from {self._path}  "
                     f"({len(self._data.get('completed_row_ids', []))} rows already done)")
                return
            except Exception:
                warn(f"Could not read {self._path} — starting fresh")
        self._data = {
            "version":           "1.1",
            "excel":             self._excel,
            "started_at":        _now_iso(),
            "last_updated":      _now_iso(),
            "completed_row_ids": [],
            "failed_row_ids":    [],
            "results":           [],
        }
        self._save()

    def _save(self):
        self._data["last_updated"] = _now_iso()
        tmp = self._path.with_suffix(".tmp")
        with open(tmp, "w") as f:
            json.dump(self._data, f, indent=2)
        tmp.replace(self._path)   # atomic rename

    # -- public API --------------------------------------------------------

    @property
    def completed_ids(self) -> set[int]:
        return set(self._data.get("completed_row_ids", []))

    @property
    def failed_ids(self) -> set[int]:
        return set(self._data.get("failed_row_ids", []))

    def record(self, result: dict):
        """Called after each server completes (thread-safe)."""
        with self._lock:
            result = dict(result)
            result["completed_at"] = _now_iso()
            row_id = result["id"]
            # Replace existing record for this ID if retrying
            self._data["results"] = [
                r for r in self._data["results"] if r["id"] != row_id
            ]
            self._data["results"].append(result)
            if result["status"] == "ok":
                ids = self._data["completed_row_ids"]
                if row_id not in ids:
                    ids.append(row_id)
                # Remove from failed if it was there before
                self._data["failed_row_ids"] = [
                    i for i in self._data["failed_row_ids"] if i != row_id
                ]
            else:
                ids = self._data["failed_row_ids"]
                if row_id not in ids:
                    ids.append(row_id)
            self._save()

    def reset(self):
        """Wipe all progress."""
        with self._lock:
            self._data = {
                "version":           "1.1",
                "excel":             self._excel,
                "started_at":        _now_iso(),
                "last_updated":      _now_iso(),
                "completed_row_ids": [],
                "failed_row_ids":    [],
                "results":           [],
            }
            self._save()
        ok(f"State reset.  {self._path} cleared.")

    def print_status(self):
        """Pretty-print the current progress without running anything."""
        hdr("CURRENT PROGRESS  (patch_state.json)")
        if not self._data.get("results"):
            info("No runs recorded yet.")
            return

        print(f"  Excel:        {self._data.get('excel', '?')}")
        print(f"  Started:      {self._data.get('started_at', '?')}")
        print(f"  Last updated: {self._data.get('last_updated', '?')}")
        done    = self._data.get("completed_row_ids", [])
        failed  = self._data.get("failed_row_ids", [])
        print(f"  Completed:    {len(done)} rows  {done}")
        print(f"  Failed:       {len(failed)} rows  {failed}")
        print()

        col = "{:<5}  {:<22}  {:<18}  {:<10}  {:<12}  {}"
        print(_c(col.format("ID", "Host", "Service", "Action", "Status", "Completed at"), Fore.WHITE))
        print("-" * 82)
        for r in sorted(self._data.get("results", []), key=lambda x: x["id"]):
            sc = _c("✔ ok", Fore.GREEN) if r["status"] == "ok" else _c("✖ error", Fore.RED)
            print(col.format(
                r["id"], r["host"][:22], r["service"][:18],
                r["action"][:10], sc, r.get("completed_at", "")[:19],
            ))


# =========================================================================
#  Config loading
# =========================================================================

def load_config(path: str) -> dict:
    p = Path(path)
    if not p.exists():
        err(f"Config file not found: {path}")
        sys.exit(1)
    with open(p) as f:
        return yaml.safe_load(f)


# =========================================================================
#  Excel loading
# =========================================================================

REQUIRED_COLS = {
    "id", "hostname / ip", "service name",
    "check command", "start command", "stop command", "mode",
}

COL_ALIASES = {
    "hostname":                       "hostname / ip",
    "ip":                             "hostname / ip",
    "host":                           "hostname / ip",
    "hostname/ip":                    "hostname / ip",
    "hostname / ip":                  "hostname / ip",
    "service":                        "service name",
    "service name":                   "service name",
    "check":                          "check command",
    "check command":                  "check command",
    "check command\n(status)":        "check command",
    "status command":                 "check command",
    "start":                          "start command",
    "start command":                  "start command",
    "stop":                           "stop command",
    "stop command":                   "stop command",
    "mode":                           "mode",
    "mode\n(round_robin / batch)":    "mode",
    "id":                             "id",
    "notes":                          "notes",
}


def load_servers(excel_path: str,
                 row_ids:    list[int] | None,
                 start_row:  int | None,
                 stop_row:   int | None) -> list[dict]:
    p = Path(excel_path)
    if not p.exists():
        err(f"Excel file not found: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(p, data_only=True)
    sheet_name = next(
        (s for s in wb.sheetnames if s.lower() != "legend"),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]
    info(f"Reading sheet '{sheet_name}' from {excel_path}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        err("Excel sheet is empty.")
        sys.exit(1)

    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    headers     = [COL_ALIASES.get(h.lower(), h.lower()) for h in raw_headers]

    missing = REQUIRED_COLS - set(headers)
    if missing:
        err(f"Missing required columns: {', '.join(sorted(missing))}")
        err(f"Found columns: {', '.join(raw_headers)}")
        sys.exit(1)

    servers = []
    for row in rows[1:]:
        rec = dict(zip(headers, row))
        if not rec.get("hostname / ip"):
            continue
        try:
            rec["id"] = int(rec.get("id") or 0)
        except (TypeError, ValueError):
            rec["id"] = 0
        mode = str(rec.get("mode") or "batch").strip().lower()
        if mode not in ("round_robin", "batch"):
            warn(f"Row {rec['id']}: unknown mode '{mode}', defaulting to 'batch'")
            mode = "batch"
        rec["mode"] = mode
        servers.append(rec)

    # -- apply filters (all are additive / AND-ed together) --------------
    if row_ids is not None:
        id_set  = set(row_ids)
        servers = [s for s in servers if s["id"] in id_set]

    if start_row is not None:
        servers = [s for s in servers if s["id"] >= start_row]

    if stop_row is not None:
        servers = [s for s in servers if s["id"] <= stop_row]

    if not servers:
        err("No servers matched the given filters.")
        sys.exit(1)

    return servers


# =========================================================================
#  SSH helpers
# =========================================================================

def _make_ssh_client(host: str, cfg: dict, verbose: bool):
    ssh_cfg = cfg.get("ssh", {})
    client  = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    kwargs = dict(
        hostname     = host,
        port         = int(ssh_cfg.get("port", 22)),
        username     = ssh_cfg.get("username", "root"),
        timeout      = float(ssh_cfg.get("connect_timeout", 15)),
        banner_timeout = 30,
        auth_timeout   = 30,
    )
    key_path = (ssh_cfg.get("private_key_path") or "").strip()
    password  = (ssh_cfg.get("password") or "").strip()
    if key_path:
        kwargs["key_filename"] = os.path.expanduser(key_path)
    elif password:
        kwargs["password"] = password

    client.connect(**kwargs)
    return client


def _run_cmd(client, cmd: str, timeout: float) -> tuple[int, str, str]:
    _, stdout, stderr = client.exec_command(cmd, timeout=timeout)
    exit_code = stdout.channel.recv_exit_status()
    return exit_code, stdout.read().decode().strip(), stderr.read().decode().strip()


# =========================================================================
#  Core patching logic — single server
# =========================================================================

def patch_server(server: dict, cfg: dict, dry_run: bool,
                 verbose: bool, logger: logging.Logger,
                 state: PatchState) -> dict:
    host      = str(server["hostname / ip"]).strip()
    service   = str(server["service name"]).strip()
    check_cmd = str(server["check command"]).strip()
    start_cmd = str(server["start command"]).strip()
    stop_cmd  = str(server["stop command"]).strip()
    row_id    = server["id"]
    timeout   = float(cfg.get("ssh", {}).get("command_timeout", 30))

    result = {"id": row_id, "host": host, "service": service,
              "action": "none", "status": "ok", "message": ""}
    label  = f"[{row_id}] {host} / {service}"

    try:
        if verbose:
            info(f"{label}  Connecting …")

        if dry_run:
            info(f"{label}  [DRY-RUN] Would check status via: {check_cmd}")
            result.update(action="dry-run", message="Dry-run — no changes made")
            logger.info("DRY-RUN  %s", label)
            state.record(result)
            return result

        if not _PARAMIKO:
            msg = "paramiko not installed – run: pip install paramiko"
            err(f"{label}  →  {msg}")
            result.update(status="error", message=msg)
            state.record(result)
            return result

        client = _make_ssh_client(host, cfg, verbose)

        # -- 1. Check current status --------------------------------------
        exit_code, _, _ = _run_cmd(client, check_cmd, timeout)

        if exit_code == 0:
            action_cmd, action_verb, done_verb = stop_cmd,  "stop",  "stopped"
        else:
            action_cmd, action_verb, done_verb = start_cmd, "start", "started"

        if verbose:
            info(f"{label}  Status exit={exit_code} → will {action_verb}")

        # -- 2. Perform toggle --------------------------------------------
        ac_exit, ac_out, ac_err = _run_cmd(client, action_cmd, timeout)
        client.close()

        if ac_exit == 0:
            ok(f"{label}  {done_verb.upper()}")
            result.update(action=action_verb,
                          message=f"Service {done_verb} successfully")
            logger.info("OK       %s  action=%s", label, action_verb)
        else:
            err_msg = ac_err or ac_out or f"exit code {ac_exit}"
            err(f"{label}  FAILED to {action_verb}: {err_msg}")
            result.update(status="error", action=action_verb,
                          message=f"Command failed: {err_msg}")
            logger.error("FAILED   %s  action=%s  error=%s",
                         label, action_verb, err_msg)

    except Exception as exc:
        err(f"{label}  ERROR: {exc}")
        result.update(status="error", message=str(exc))
        logger.error("ERROR    %s  %s", label, exc)

    state.record(result)
    return result


# =========================================================================
#  Orchestration
# =========================================================================

def run_round_robin(servers: list[dict], cfg: dict, dry_run: bool,
                    verbose: bool, logger: logging.Logger,
                    state: PatchState) -> list[dict]:
    delay   = float(cfg.get("patching", {}).get("round_robin_delay", 5))
    results = []
    total   = len(servers)
    for idx, server in enumerate(servers, start=1):
        info(f"Round-robin  {idx}/{total}  →  {server['hostname / ip']}")
        r = patch_server(server, cfg, dry_run, verbose, logger, state)
        results.append(r)
        if idx < total and not dry_run:
            info(f"Waiting {delay}s before next server …")
            time.sleep(delay)
    return results


def run_batch(servers: list[dict], cfg: dict, dry_run: bool,
              verbose: bool, logger: logging.Logger,
              state: PatchState) -> list[dict]:
    max_workers = int(cfg.get("patching", {}).get("batch_max_workers", 10))
    results     = []
    with ThreadPoolExecutor(max_workers=min(max_workers, len(servers))) as pool:
        futures = {
            pool.submit(patch_server, s, cfg, dry_run, verbose, logger, state): s
            for s in servers
        }
        for fut in as_completed(futures):
            results.append(fut.result())
    results.sort(key=lambda r: r["id"])
    return results


# =========================================================================
#  Summary table
# =========================================================================

def print_summary(results: list[dict], skipped: list[int]):
    hdr("SUMMARY")
    col = "{:<5}  {:<22}  {:<20}  {:<10}  {}"
    print(_c(col.format("ID", "Host", "Service", "Action", "Result"), Fore.WHITE))
    print("-" * 78)
    for r in results:
        sc  = (_c("✔ OK", Fore.GREEN) if r["status"] == "ok"
               else _c("✖ ERROR", Fore.RED))
        msg = r["message"][:45] + "…" if len(r["message"]) > 45 else r["message"]
        print(col.format(r["id"], r["host"][:22], r["service"][:20],
                         r["action"][:10], f"{sc}  {msg}"))

    total  = len(results)
    errors = sum(1 for r in results if r["status"] != "ok")
    print()
    print(_c(f"  Processed: {total}   OK: {total - errors}   "
             f"Errors: {errors}   Skipped (already done): {len(skipped)}", Fore.CYAN))
    if skipped:
        info(f"Skipped row IDs: {skipped}  (use --reset-state to re-run them)")


# =========================================================================
#  CLI
# =========================================================================

def parse_args():
    p = argparse.ArgumentParser(
        description="Linux Server Patching Agent — toggle services via SSH",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    _here = Path(__file__).parent
    p.add_argument("--excel",   default=str(_here.parent / "servers_template.xlsx"),
                   help="Path to the Excel file (default: ../servers_template.xlsx)")
    p.add_argument("--config",  default=str(_here / "config.yaml"),
                   help="Path to the YAML config file (default: config.yaml)")

    # -- row selection --------------------------------------------------
    sel = p.add_argument_group("Row selection (can be combined)")
    sel.add_argument("--rows",      default=None,
                     help="Comma-separated specific row IDs, e.g. --rows 1,3,5")
    sel.add_argument("--start-row", type=int, default=None, metavar="N",
                     help="Process rows with ID >= N  (inclusive)")
    sel.add_argument("--stop-row",  type=int, default=None, metavar="N",
                     help="Process rows with ID <= N  (inclusive)")

    # -- progress / state -----------------------------------------------
    st = p.add_argument_group("Progress / state")
    st.add_argument("--resume",      action="store_true",
                    help="Skip rows already marked completed in patch_state.json")
    st.add_argument("--status",      action="store_true",
                    help="Print current progress from patch_state.json and exit")
    st.add_argument("--reset-state", action="store_true",
                    help="Clear patch_state.json and start fresh")
    st.add_argument("--state-file",  default=STATE_FILE, metavar="FILE",
                    help=f"Path to the state file (default: {STATE_FILE})")

    # -- misc -----------------------------------------------------------
    p.add_argument("--dry-run",       action="store_true",
                   help="Show what would happen without touching any server")
    p.add_argument("--verbose", "-v", action="store_true",
                   help="Show extra SSH connection details")
    p.add_argument("--mode-override", choices=["round_robin", "batch"], default=None,
                   help="Force all servers to this mode, ignoring the Excel column")
    return p.parse_args()


def main():
    print(_c(BANNER, Fore.CYAN))
    args = parse_args()

    # -- parse row_ids ------------------------------------------------
    row_ids = None
    if args.rows:
        try:
            row_ids = [int(x.strip()) for x in args.rows.split(",")]
        except ValueError:
            err("--rows must be comma-separated integers, e.g. --rows 1,3,5")
            sys.exit(1)

    # -- state file ----------------------------------------------------
    state = PatchState(excel=args.excel, state_file=args.state_file)

    if args.reset_state:
        state.reset()
        sys.exit(0)

    if args.status:
        state.print_status()
        sys.exit(0)

    # -- load config & servers -----------------------------------------
    cfg     = load_config(args.config)
    servers = load_servers(args.excel, row_ids, args.start_row, args.stop_row)
    log_file = cfg.get("patching", {}).get("log_file", "patch_agent.log")
    logger   = setup_logging(log_file)

    if args.mode_override:
        for s in servers:
            s["mode"] = args.mode_override

    # -- --resume: drop rows already done -----------------------------
    skipped = []
    if args.resume:
        done = state.completed_ids
        skipped = sorted(s["id"] for s in servers if s["id"] in done)
        servers  = [s for s in servers if s["id"] not in done]
        if skipped:
            info(f"Resuming — skipping {len(skipped)} already-completed rows: {skipped}")
        if not servers:
            ok("Nothing left to do — all selected rows are already complete.")
            state.print_status()
            sys.exit(0)

    rr_servers    = [s for s in servers if s["mode"] == "round_robin"]
    batch_servers = [s for s in servers if s["mode"] == "batch"]

    # -- print plan before starting ------------------------------------
    range_desc = ""
    if args.start_row or args.stop_row:
        lo = args.start_row or "start"
        hi = args.stop_row  or "end"
        range_desc = f"  row range: {lo}–{hi}"
    info(f"Servers to process:  {len(servers)} total  "
         f"({len(rr_servers)} round-robin, {len(batch_servers)} batch){range_desc}")

    if args.dry_run:
        warn("DRY-RUN mode — no changes will be made")
    if not _PARAMIKO:
        warn("paramiko is not installed.  Install it with:  pip install paramiko")
        warn("SSH operations will fail until paramiko is available.")

    all_results = []

    # -- Phase 1: round-robin ------------------------------------------
    if rr_servers:
        hdr(f"PHASE 1 — Round-Robin  ({len(rr_servers)} servers, sequential)")
        all_results += run_round_robin(rr_servers, cfg, args.dry_run,
                                       args.verbose, logger, state)

    # -- Phase 2: batch ------------------------------------------------
    if batch_servers:
        hdr(f"PHASE 2 — Batch  ({len(batch_servers)} servers, parallel)")
        all_results += run_batch(batch_servers, cfg, args.dry_run,
                                 args.verbose, logger, state)

    print_summary(all_results, skipped)

    if log_file:
        info(f"Full log:    {log_file}")
    info(f"State file:  {args.state_file}")

    errors = sum(1 for r in all_results if r["status"] != "ok")
    sys.exit(1 if errors else 0)


if __name__ == "__main__":
    main()
